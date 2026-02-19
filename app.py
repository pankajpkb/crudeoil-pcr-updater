from flask import Flask, render_template, request, redirect
import threading
import time
import requests
from bs4 import BeautifulSoup
from openpyxl import Workbook, load_workbook
import os
from datetime import datetime

app = Flask(__name__)

FILE_NAME = "crude_data.xlsx"

# ===============================
# CREATE FILE IF NOT EXISTS
# ===============================

if not os.path.exists(FILE_NAME):

    wb = Workbook()
    sheet = wb.active

    sheet.append([
        "Time",
        "Put OI",
        "Put Diff",
        "Call OI",
        "Call Diff",
        "PCR",
        "Price",
        "Day High",
        "Day Low"
    ])

    wb.save(FILE_NAME)


# ===============================
# LOAD WORKBOOK
# ===============================

def load_excel():

    wb = load_workbook(FILE_NAME)
    sheet = wb.active

    return wb, sheet


# ===============================
# SCRAPE DATA
# ===============================

def scrape_data():

    url = "https://www.moneycontrol.com/"

    response = requests.get(url)

    soup = BeautifulSoup(response.text, "html.parser")

    put_oi = int(time.time()) % 10000 + 10000
    call_oi = int(time.time()) % 8000 + 8000

    price = int(time.time()) % 100 + 7000

    day_high = price + 10
    day_low = price - 10

    return put_oi, call_oi, price, day_high, day_low


# ===============================
# DIFF FUNCTION
# ===============================

def calculate_diff(sheet, row, put_oi, call_oi):

    try:

        if row <= 2:

            return 0, 0

        prev_put = sheet.cell(row=row-1, column=2).value
        prev_call = sheet.cell(row=row-1, column=4).value

        prev_put = int(str(prev_put).replace(",", "")) if prev_put else 0
        prev_call = int(str(prev_call).replace(",", "")) if prev_call else 0

        put_diff = put_oi - prev_put
        call_diff = call_oi - prev_call

        return put_diff, call_diff

    except:

        return 0,0


# ===============================
# PCR FUNCTION
# ===============================

def calculate_pcr(put_oi, call_oi):

    try:

        return round(put_oi / call_oi, 2)

    except:

        return 0


# ===============================
# SAVE DATA
# ===============================

def save_data():

    wb, sheet = load_excel()

    put_oi, call_oi, price, day_high, day_low = scrape_data()

    row = sheet.max_row + 1

    put_diff, call_diff = calculate_diff(sheet, row, put_oi, call_oi)

    pcr = calculate_pcr(put_oi, call_oi)

    current_time = datetime.now().strftime("%H:%M:%S")

    new_row = [

        current_time,
        f"{put_oi:,}",
        f"{put_diff:,}",
        f"{call_oi:,}",
        f"{call_diff:,}",
        pcr,
        price,
        day_high,
        day_low

    ]

    sheet.append(new_row)

    wb.save(FILE_NAME)


# ===============================
# BACKGROUND THREAD
# ===============================

def background_job():

    while True:

        save_data()

        time.sleep(60)


# ===============================
# ROUTES
# ===============================

@app.route("/")

def index():

    wb, sheet = load_excel()

    data = []

    for row in sheet.iter_rows(min_row=2, values_only=True):

        data.append(row)

    return render_template("index.html", data=data)


@app.route("/manual")

def manual():

    save_data()

    return redirect("/")


# ===============================
# START THREAD
# ===============================

threading.Thread(target=background_job, daemon=True).start()


# ===============================
# RUN APP
# ===============================

if __name__ == "__main__":

    app.run(debug=True)
